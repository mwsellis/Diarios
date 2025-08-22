import os
import shutil
from openpyxl import load_workbook

def processar_arquivos(diretorio):
    modelo_diario = os.path.join(diretorio, "2025 DIARIO CEMAT REGULAR.xlsx")
    if not os.path.exists(modelo_diario):
        print("Arquivo modelo não encontrado!")
        return

    for arquivo in os.listdir(diretorio):
        if arquivo.startswith("RelAlunosMatPTurma") and arquivo.endswith(".xlsx"):
            partes = arquivo.split()
            if len(partes) < 3:
                continue  # Garante que há informações suficientes no nome do arquivo

            turma = partes[-2]  # Última palavra antes da turma
            turno = partes[-1].split('.')[0]  # Última palavra numérica antes da extensão

            novo_nome = f"2025 DIARIO CEMAT REGULAR {turma}.xlsx"
            novo_caminho = os.path.join(diretorio, novo_nome)

            shutil.copy(modelo_diario, novo_caminho)

            wb_origem = load_workbook(os.path.join(diretorio, arquivo))
            wb_destino = load_workbook(novo_caminho)

            if "RelAlunosMatPTurma" in wb_origem.sheetnames and "ALUNOS" in wb_destino.sheetnames:
                ws_origem = wb_origem["RelAlunosMatPTurma"]
                ws_destino = wb_destino["ALUNOS"]

                for linha in range(6, 51):
                    ws_destino[f"B{linha}"] = ws_origem[f"B{linha}"].value
                    ws_destino[f"C{linha}"] = ws_origem[f"C{linha}"].value

                # Preenchendo B2 com a turma e C2 com o turno
                ws_destino["B2"] = turma
                ws_destino["C2"] = turno

                for aba in ["1o Trimestre", "2o Trimestre", "3o Trimestre", "Resultados"]:
                    if aba in wb_destino.sheetnames:
                        ws = wb_destino[aba]
                        ws.protection.sheet = True

                wb_destino.save(novo_caminho)
                print(f"Processado: {novo_nome}")
            else:
                print(f"Planilha 'ALUNOS' não encontrada em {arquivo}")

            wb_origem.close()
            wb_destino.close()

# Defina o caminho do diretório onde estão os arquivos
diretorio_raiz = os.path.abspath("Diários 2025 - 2TRIM/Matriculados")
processar_arquivos(diretorio_raiz)
