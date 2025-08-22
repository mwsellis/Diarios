import os
import shutil
import datetime
from openpyxl import load_workbook

def processar_arquivos(diretorio):
    modelo_diario = os.path.join(diretorio, "Modelo 3o trimestre - Impressão.xlsx")
    ws_contador = 0
    timeInicio = datetime.datetime.now()

    print("Inicio do processamento")
    if not os.path.exists(modelo_diario):
        print("Arquivo modelo não encontrado!")
        return

    for arquivo in os.listdir(diretorio):
        if arquivo.startswith("RelAlunosMatPTurma") and arquivo.endswith(".xlsx"):
            partes = arquivo.split()
            if len(partes) < 3:
                continue

            turma = partes[-2]
            turno = partes[-1].split('.')[0]

            novo_nome = f"3o trimestre - Impressão - {turma}.xlsx"
            novo_caminho = os.path.join(diretorio, novo_nome)

            shutil.copy(modelo_diario, novo_caminho)

            ws_contador +=1

            wb_origem = load_workbook(os.path.join(diretorio, arquivo))
            wb_destino = load_workbook(novo_caminho)

            if "RelAlunosMatPTurma" in wb_origem.sheetnames and "NOTAS-FREQ" in wb_destino.sheetnames:
                ws_origem = wb_origem["RelAlunosMatPTurma"]
                ws_destino = wb_destino["NOTAS-FREQ"]
                ws_conteudo = wb_destino["CONTEÚDOS"]

                for linha in range(6, 36):
                    ws_destino[f"B{linha + 3}"] = ws_origem[f"B{linha}"].value
                    ws_destino[f"C{linha + 3}"] = ws_origem[f"C{linha}"].value

                for linha in range(36, 55):
                    ws_destino[f"B{linha + 13}"] = ws_origem[f"B{linha}"].value
                    ws_destino[f"C{linha + 13}"] = ws_origem[f"C{linha}"].value

                ws_destino["C5"] = turma
                ws_destino["C6"] = turno
                ws_conteudo["B6"] = turma + " - " + turno

                wb_destino.save(novo_caminho)
                print(f"Processado: {novo_nome}")
            else:
                print(f"Planilha 'ALUNOS' não encontrada em {arquivo}")

            wb_origem.close()
            wb_destino.close()

    print("Fim do processamento em: " + str(datetime.datetime.now() - timeInicio))
    print("Total de planilhas geradas:")
    print(ws_contador)

diretorio_raiz = os.path.abspath("Diários 2025 - 3TRIM/Matriculados")
processar_arquivos(diretorio_raiz)
